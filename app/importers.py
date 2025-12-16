from __future__ import annotations

from datetime import datetime
import hashlib
from pathlib import Path
from typing import Any, Iterable, Sequence
import tempfile
import shutil

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlmodel import Session, select

from . import models
from .database import get_session

def _excel_col_letter(n: int) -> str:
    """1-based column number -> Excel letter (A..Z, AA..)."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _safe_cell(v: Any):
    """Make values JSON-safe and stable for UI."""
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            pass
    try:
        if isinstance(v, (int, float, str, bool)):
            return v
    except Exception:
        pass
    return str(v)


def _detect_header_row(ws) -> int | None:
    """Heuristic: pick row (within first 25) with most non-empty strings."""
    best_row = None
    best_score = 0
    for r in range(1, min(25, ws.max_row or 1) + 1):
        vals = [c.value for c in ws[r]]
        strings = [v for v in vals if isinstance(v, str) and v.strip()]
        score = len(strings)
        if score > best_score:
            best_score = score
            best_row = r
    if best_score >= 3:
        return best_row
    return None


def store_workbook_raw(path: Path, session: Session, filename_override: str | None = None) -> dict:
    """Store *all* sheets/rows from an .xlsx into ExcelWorkbook/ExcelSheet/ExcelRow."""
    if not path.exists():
        raise HTTPException(status_code=400, detail=f"File not found: {path}")

    blob = path.read_bytes()
    sha = hashlib.sha256(blob).hexdigest()
    filename = filename_override or path.name
    now = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

    existing = session.exec(
        select(models.ExcelWorkbook).where(models.ExcelWorkbook.sha256 == sha)
    ).first()
    if existing:
        return {"workbook_id": existing.id, "filename": existing.filename, "sha256": sha, "deduped": True}

    wb = openpyxl.load_workbook(path, data_only=False)
    book = models.ExcelWorkbook(filename=filename, sha256=sha, imported_at=now)
    session.add(book)
    session.flush()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = _detect_header_row(ws)

        columns: list[str] = []
        if ws.max_column:
            if header_row:
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=header_row, column=c).value
                    key = (str(v).strip() if isinstance(v, str) and str(v).strip() else _excel_col_letter(c))
                    columns.append(key)
            else:
                columns = [_excel_col_letter(c) for c in range(1, ws.max_column + 1)]

        sheet = models.ExcelSheet(
            workbook_id=book.id,
            name=sheet_name,
            header_row=header_row,
            max_row=int(ws.max_row or 0),
            max_col=int(ws.max_column or 0),
            columns=columns,
        )
        session.add(sheet)
        session.flush()
        for r in range(1, (ws.max_row or 0) + 1):
            values = [ws.cell(row=r, column=c).value for c in range(1, (ws.max_column or 0) + 1)]
            if all(v is None for v in values):
                continue

            row_data: dict[str, Any] = {}
            for c, v in enumerate(values, start=1):
                if v is None:
                    continue
                col_key = columns[c - 1] if (c - 1) < len(columns) else _excel_col_letter(c)
                row_data[col_key] = _safe_cell(v)

            session.add(models.ExcelRow(sheet_id=sheet.id, row_index=r, data=row_data))

    session.commit()
    return {"workbook_id": book.id, "filename": book.filename, "sha256": sha, "deduped": False}

class Enum1ImportRequest(BaseModel):
    file_path: str
    sheet_name: str = "Enum-1"


class Enum1ImportResult(BaseModel):
    rows_ingested: int
    rows_skipped: int
    message: str


class SheetImportRequest(BaseModel):
    file_path: str
    sheet_name: str | None = None


def _get_or_create(session: Session, model, defaults: dict | None = None, **filters):
    stmt = select(model).filter_by(**filters)
    obj = session.exec(stmt).first()
    if obj:
        return obj
    obj = model(**filters, **(defaults or {}))
    session.add(obj)
    session.flush()
    session.refresh(obj)
    return obj


def _apply_updates(target, **fields):
    for key, value in fields.items():
        if value is None:
            continue
        setattr(target, key, value)


def _to_str(value) -> str | None:
    """Convert Excel values (which can be floats) to strings safely."""
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)):
        # Convert float to int if it's a whole number, then to string
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def _get_or_create_by_code(
    session: Session,
    model,
    code_value: str,
    defaults: dict | None = None,
):
    """
    Lookup strictly by unique 'code' column, to avoid UNIQUE constraint failures.
    If creating new, defaults MUST contain any NOT NULL foreign keys for this model.
    """
    stmt = select(model).where(model.code == code_value)
    obj = session.exec(stmt).first()
    if obj:
        return obj

    obj = model(code=code_value, **(defaults or {}))
    session.add(obj)
    session.flush()      # flush is safe only if defaults satisfy NOT NULL constraints
    session.refresh(obj)
    return obj


def _find_header_row(rows: Iterable[Sequence], required_columns: list[str]) -> tuple[dict[str, int], Iterable[Sequence]]:
    rows = iter(rows)
    for row in rows:
        header_index = {h: i for i, h in enumerate(row) if isinstance(h, str)}
        if all(col in header_index for col in required_columns):
            return header_index, rows
    raise HTTPException(status_code=400, detail="Could not find header row with required columns")


def import_enum1(path: Path, sheet_name: str, session: Session) -> Enum1ImportResult:
    if not path.exists():
        raise HTTPException(status_code=400, detail=f"File not found: {path}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to open workbook: {str(e)}")
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    ws = wb[sheet_name]
    header_index, rows = _find_header_row(
        ws.iter_rows(values_only=True),
        ["Component ID", "Component Type", "Region", "District"],
    )

    def col(name: str, row: tuple):
        idx = header_index.get(name)
        return row[idx] if idx is not None else None

    ingested = skipped = 0

    for row in rows:
        component_code = col("Component ID", row)
        component_type = col("Component Type", row)
        region_name = col("Region", row)
        district_name = col("District", row)

        landmark_code = col("Landmark ID", row)
        pole_code = col("Pole Location", row)
        jb_code = col("JB ID", row)

        if not component_code or not component_type or not region_name or not district_name:
            skipped += 1
            continue

        region = _get_or_create(session, models.Region, name=str(region_name))
        district = _get_or_create(session, models.District, name=str(district_name), region_id=region.id)

        # -------- Landmark (code is unique, district_id+region_id are NOT NULL) --------
        landmark = None
        if landmark_code:
            landmark = _get_or_create_by_code(
                session,
                models.Landmark,
                str(landmark_code),
                defaults={"district_id": district.id, "region_id": region.id},
            )
            _apply_updates(landmark, district_id=district.id, region_id=region.id)
            session.add(landmark)

        # If we have pole but no landmark, skip row
        if pole_code and not landmark:
            skipped += 1
            continue

        # -------- Pole (code unique, landmark_id/district_id/region_id are NOT NULL) --------
        pole = None
        if pole_code and landmark:
            pole = _get_or_create_by_code(
                session,
                models.Pole,
                str(pole_code),
                defaults={
                    "landmark_id": landmark.id,
                    "district_id": district.id,
                    "region_id": region.id,
                },
            )
            _apply_updates(
                pole,
                landmark_id=landmark.id,
                district_id=district.id,
                region_id=region.id,
            )
            session.add(pole)

        # If we have JB but no pole, skip row
        if jb_code and not pole:
            skipped += 1
            continue

        # -------- JunctionBox (code unique, pole_id/landmark_id/district_id/region_id are NOT NULL) --------
        jb = None
        if jb_code and pole:
            jb = _get_or_create_by_code(
                session,
                models.JunctionBox,
                str(jb_code),
                defaults={
                    "pole_id": pole.id,
                    "landmark_id": landmark.id,
                    "district_id": district.id,
                    "region_id": region.id,
                },
            )
            _apply_updates(
                jb,
                pole_id=pole.id,
                landmark_id=landmark.id,
                district_id=district.id,
                region_id=region.id,
            )
            session.add(jb)

        payload = models.Component(
            component_code=_to_str(component_code),
            component_type=_to_str(component_type),

            connected_to_code=_to_str(col("Connected To (Component ID)", row)),
            model=_to_str(col("Model/ Specific Device", row)),
            serial=_to_str(col("Manufacturer Serial Number", row)),
            firmware=_to_str(col("Firmware/ Software Version", row)),
            os=_to_str(col("Operating System (if applicable)", row)),
            licenses=_to_str(col("Software Licenses (if applicable)", row)),

            pole_id=pole.id if pole else None,
            jb_id=jb.id if jb else None,
            landmark_id=landmark.id if landmark else None,
            district_id=district.id,
            region_id=region.id,

            project_phase=col("Project Phase", row),
            lat=col("Latitude", row),
            lng=col("Longitude", row),
            landmark_name=_to_str(col("Landmark", row)),
            frs_camera=_to_str(col("FRS Camera", row)),
            power_req=_to_str(col("Power Source/ Requirements", row)),

            local_if_name=_to_str(col("Local Interface Name/Port", row)),
            local_if_ip=_to_str(col("Local Interface IP Address", row)),
            remote_if_name=_to_str(col("Remote Interface Name/Port", row)),
            remote_if_ip=_to_str(col("Remote Interface IP Address", row)),

            cable_id=_to_str(col("Cable ID", row)),
            physical_link_type=_to_str(col("Physical Link Type", row)),
            logical_link_type=_to_str(col("Logical Link Type (Overall Network Segment)", row)),
            segment_type=_to_str(col("Segment Type (Connectivity Model)", row)),
            segment_switches=_to_str(col("Segment Structure - Switches", row)),
            segment_junctions=_to_str(col("Segment Structure - Junctions", row)),
            segment_instance_no=_to_str(col("Segment Structure - Instance Number", row)),
            fiber_core_usage=_to_str(col("Fiber Core Usage (if OFC)", row)),

            proposed_vlan=_to_str(col("Proposed VLAN ID", row)),
            proposed_subnet=_to_str(col("Proposed Subnet (CIDR)", row)),
            ip_assignment=_to_str(col("IP Assignment Method", row)),
            video_priority=_to_str(col("Video-High-Priority (EF)", row)),
            security_zone=_to_str(col("Security Zone/Firewall Zone", row)),

            last_config_change=_to_str(col("Last Configuration Change Date", row)),
            last_config_backup=_to_str(col("Last Configuration Backup Date", row)),
            maintenance_schedule=_to_str(col("Maintenance Schedule", row)),
            last_maintenance=_to_str(col("Last Maintenance Date", row)),
            monitoring_tool=_to_str(col("Monitoring Status/Tool", row)),

            network_provider=_to_str(col("Network Provider", row)),
            static_router_ip=_to_str(col("Static IP Of router ", row)),
            landline_number=_to_str(col("Landline Number", row)),
            termination_type=_to_str(col("Termination Type(Port Forwarding /VPN)", row)),

            router1=_to_str(col("Router 1", row)),
            router2=_to_str(col("Router 2", row)),
            http_port=_to_str(col("HTTP Port ", row)),
            rtsp_port=_to_str(col("RTSP Port", row)),
        )

        existing = session.exec(
            select(models.Component).where(models.Component.component_code == payload.component_code)
        ).first()

        if existing:
            for k, v in payload.dict(exclude_unset=True).items():
                setattr(existing, k, v)
            session.add(existing)
        else:
            session.add(payload)

        ingested += 1

    session.commit()
    return Enum1ImportResult(rows_ingested=ingested, rows_skipped=skipped, message="Import finished")


def import_ip_schema_data(path: Path, sheet_name: str, session: Session) -> Enum1ImportResult:
    """Import poles and landmarks from IP SCHEMA.xlsx"""
    if not path.exists():
        raise HTTPException(status_code=400, detail=f"File not found: {path}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to open workbook: {str(e)}")
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    # Find the actual header row (row 5 in the file, 0-indexed)
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == 'Sr.':
            header_row_idx = i
            break
    
    if header_row_idx is None:
        raise HTTPException(status_code=400, detail="Could not find header row with 'Sr.' column")
    
    header_row = rows[header_row_idx]
    header_index = {h: i for i, h in enumerate(header_row) if h is not None}
    
    def col(name: str, row: tuple):
        idx = header_index.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    ingested = skipped = 0

    for row in rows[header_row_idx + 1:]:
        landmark_code = col("Landmark ID", row)
        pole_code = col("Pole Location", row)
        region_name = col("Region", row)
        district_name = col("District", row)
        jb_code = col("JB ID", row)

        if not pole_code or not region_name or not district_name or not landmark_code:
            skipped += 1
            continue

        region = _get_or_create(session, models.Region, name=str(region_name))
        district = _get_or_create(session, models.District, name=str(district_name), region_id=region.id)

        # landmark (NOT NULL district_id/region_id)
        landmark = _get_or_create_by_code(
            session,
            models.Landmark,
            str(landmark_code),
            defaults={"district_id": district.id, "region_id": region.id},
        )
        _apply_updates(landmark, district_id=district.id, region_id=region.id, 
                       name=col("Landmark", row), lat=col("Latitude", row), lng=col("Longitude", row))
        session.add(landmark)

        # pole (NOT NULL landmark_id/district_id/region_id)
        pole = _get_or_create_by_code(
            session,
            models.Pole,
            str(pole_code),
            defaults={
                "landmark_id": landmark.id,
                "district_id": district.id,
                "region_id": region.id,
            },
        )
        _apply_updates(pole, landmark_id=landmark.id, district_id=district.id, region_id=region.id,
                       location_name=col("Landmark", row), lat=col("Latitude", row), lng=col("Longitude", row))
        session.add(pole)

        jb_code_val = col("JB ID", row)
        if jb_code_val:
            jb = _get_or_create_by_code(
                session,
                models.JunctionBox,
                str(jb_code_val),
                defaults={
                    "pole_id": pole.id,
                    "landmark_id": landmark.id,
                    "district_id": district.id,
                    "region_id": region.id,
                },
            )
            _apply_updates(
                jb,
                pole_id=pole.id,
                landmark_id=landmark.id,
                district_id=district.id,
                region_id=region.id,
                lat=col("Latitude", row),
                lng=col("Longitude", row),
            )
            session.add(jb)

        ingested += 1

    session.commit()
    return Enum1ImportResult(rows_ingested=ingested, rows_skipped=skipped, message="IP schema data import finished")


def import_field_device_jbs(path: Path, sheet_name: str, session: Session) -> Enum1ImportResult:
    """Import junction boxes from IP SCHEMA.xlsx"""
    if not path.exists():
        raise HTTPException(status_code=400, detail=f"File not found: {path}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to open workbook: {str(e)}")
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    # Find the actual header row
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == 'Sr.':
            header_row_idx = i
            break
    
    if header_row_idx is None:
        raise HTTPException(status_code=400, detail="Could not find header row with 'Sr.' column")
    
    header_row = rows[header_row_idx]
    header_index = {h: i for i, h in enumerate(header_row) if h is not None}
    
    def col(name: str, row: tuple):
        idx = header_index.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    ingested = skipped = 0

    for row in rows[header_row_idx + 1:]:
        jb_code = col("JB ID", row)
        pole_code = col("Pole Location", row)
        region_name = col("Region", row)
        district_name = col("District", row)
        landmark_code = col("Landmark ID", row)

        if not jb_code or not pole_code or not region_name or not district_name or not landmark_code:
            skipped += 1
            continue

        region = _get_or_create(session, models.Region, name=str(region_name))
        district = _get_or_create(session, models.District, name=str(district_name), region_id=region.id)

        landmark = _get_or_create_by_code(
            session,
            models.Landmark,
            str(landmark_code),
            defaults={"district_id": district.id, "region_id": region.id},
        )
        _apply_updates(landmark, district_id=district.id, region_id=region.id, 
                       name=col("Landmark", row), lat=col("Latitude", row), lng=col("Longitude", row))
        session.add(landmark)

        pole = _get_or_create_by_code(
            session,
            models.Pole,
            str(pole_code),
            defaults={
                "landmark_id": landmark.id,
                "district_id": district.id,
                "region_id": region.id,
            },
        )
        _apply_updates(pole, landmark_id=landmark.id, district_id=district.id, region_id=region.id,
                       location_name=col("Landmark", row), lat=col("Latitude", row), lng=col("Longitude", row))
        session.add(pole)

        jb = _get_or_create_by_code(
            session,
            models.JunctionBox,
            str(jb_code),
            defaults={
                "pole_id": pole.id,
                "landmark_id": landmark.id,
                "district_id": district.id,
                "region_id": region.id,
            },
        )
        _apply_updates(
            jb,
            pole_id=pole.id,
            landmark_id=landmark.id,
            district_id=district.id,
            region_id=region.id,
            lat=col("Latitude", row),
            lng=col("Longitude", row),
        )
        session.add(jb)

        ingested += 1

    session.commit()
    return Enum1ImportResult(rows_ingested=ingested, rows_skipped=skipped, message="IP schema junction-box import finished")


router = APIRouter(prefix="/import", tags=["Import"])


@router.get("/sheets")
def list_sheets(file_path: str):
    path = Path(file_path)
    if not path.exists():
        raise HTTPException(status_code=400, detail=f"File not found: {file_path}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        return {"sheets": wb.sheetnames, "file_path": file_path}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to open workbook: {str(e)}")


@router.post("/enum1", response_model=Enum1ImportResult)
def import_enum1_endpoint(request: Enum1ImportRequest, session: Session = Depends(get_session)):
    return import_enum1(Path(request.file_path), request.sheet_name, session)


@router.post("/ip-schema/poles", response_model=Enum1ImportResult)
def import_ip_schema_poles(request: SheetImportRequest, session: Session = Depends(get_session)):
    sheet = request.sheet_name or "Field Device Details - Poles"
    return import_ip_schema_data(Path(request.file_path), sheet, session)


@router.post("/ip-schema/jbs", response_model=Enum1ImportResult)
def import_ip_schema_jbs(request: SheetImportRequest, session: Session = Depends(get_session)):
    sheet = request.sheet_name or "Field Device Details - JB"
    return import_field_device_jbs(Path(request.file_path), sheet, session)


def import_credentials(path: Path, sheet_name: str, session: Session) -> Enum1ImportResult:
    """Import credentials from PHASE 1 CREDENTIALS.xlsx"""
    if not path.exists():
        raise HTTPException(status_code=400, detail=f"File not found: {path}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to open workbook: {str(e)}")
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    # Row 0 has main headers, Row 1 has location subheaders
    if len(rows) < 3:
        raise HTTPException(status_code=400, detail="Insufficient rows in sheet")
    
    main_header = rows[0]
    location_header = rows[1]
    
    # Create combined header index from both rows
    header_index = {}
    for i, (main, location) in enumerate(zip(main_header, location_header)):
        if main:
            header_index[main] = i
        if location:
            header_index[location] = i
    
    def col(name: str, row: tuple):
        idx = header_index.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    ingested = skipped = 0

    # Data starts from row 2
    for row in rows[2:]:
        s_no = col("S NO", row)
        appliance = col("APPLIANCE", row)
        username = col(" USER ID", row) or col("OS USER ID", row)
        password = col(" PASSWORD", row) or col("OS PASSWORD", row)
        ip_address = col("IP", row)
        hostname = col("HOSTNAME", row)
        access_type = col("SNMP VERSION", row)
        snmp_community = col("SNMP COMMUNITY STRING 1", row)
        snmp_server = col("SNMP SERVER IP 1", row)
        snmp_trap = col("SNMP TRAP ENABLED", row)
        region = col("Region", row)
        location_name = col("Location", row)
        
        # Check if row has ANY useful data (credentials OR SNMP config)
        has_credentials = any([appliance, username, password, ip_address, hostname])
        has_snmp = any([access_type, snmp_community, snmp_server, snmp_trap])
        
        if not (has_credentials or has_snmp):
            skipped += 1
            continue
        
        # Create unique identifier for credential (prefer IP/hostname, fallback to appliance or S NO)
        unique_id = hostname or ip_address or appliance or snmp_server or s_no
        if not unique_id:
            skipped += 1
            continue

        credential_code = f"{sheet_name}-{unique_id}"
        
        credential = session.exec(
            select(models.Credential).where(models.Credential.component_code == credential_code)
        ).first()

        notes = ""
        if appliance:
            notes = f"Appliance: {appliance}"
        if region:
            notes += (", " if notes else "") + f"Region: {region}"
        if location_name:
            notes += (", " if notes else "") + f"Location: {location_name}"
        if snmp_community:
            notes += (", " if notes else "") + f"SNMP Community: {snmp_community}"
        if snmp_server:
            notes += (", " if notes else "") + f"SNMP Server: {snmp_server}"

        payload_data = {
            "component_code": _to_str(credential_code),
            "username": _to_str(username),
            "password": _to_str(password),
            "ip_address": _to_str(ip_address),
            "port": _to_str(col("Port", row)),
            "access_type": _to_str(access_type) or _to_str(snmp_community),
            "notes": _to_str(notes) if notes else None,
            "last_updated": _to_str(col("Last Updated", row)),
        }

        if credential:
            for k, v in payload_data.items():
                if v is not None:
                    setattr(credential, k, v)
            session.add(credential)
        else:
            session.add(models.Credential(**payload_data))

        ingested += 1

    session.commit()
    return Enum1ImportResult(rows_ingested=ingested, rows_skipped=skipped, message=f"Credentials import from {sheet_name} finished")


@router.post("/credentials", response_model=Enum1ImportResult)
def import_credentials_endpoint(request: SheetImportRequest, session: Session = Depends(get_session)):
    sheet = request.sheet_name or "Sheet1"
    return import_credentials(Path(request.file_path), sheet, session)


class ImportAllRequest(BaseModel):
    enum1_path: str
    enum1_sheet: str | None = None
    ip_path: str
    ip_poles_sheet: str | None = None
    ip_jbs_sheet: str | None = None
    credentials_path: str
    credentials_sheet: str | None = None


@router.post("/all")
def import_all_endpoint(request: ImportAllRequest, session: Session = Depends(get_session)):
    """Run all imports in sequence and return a summary."""
    results = {}

    # Enum-1 (JKP Network Design Draft)
    try:
        sheet = request.enum1_sheet or "Enum-1"
        r = import_enum1(Path(request.enum1_path), sheet, session)
        results["enum1"] = r.dict()
    except HTTPException as e:
        results["enum1"] = {"error": str(e.detail)}

    # IP Schema - Poles
    try:
        sheet = request.ip_poles_sheet or "Field Device Details - Poles"
        r = import_ip_schema_data(Path(request.ip_path), sheet, session)
        results["ip_poles"] = r.dict()
    except HTTPException as e:
        results["ip_poles"] = {"error": str(e.detail)}

    # IP Schema - JBs
    try:
        sheet = request.ip_jbs_sheet or "Field Device Details - JB"
        r = import_field_device_jbs(Path(request.ip_path), sheet, session)
        results["ip_jbs"] = r.dict()
    except HTTPException as e:
        results["ip_jbs"] = {"error": str(e.detail)}

    # Credentials
    try:
        sheet = request.credentials_sheet or "Sheet1"
        r = import_credentials(Path(request.credentials_path), sheet, session)
        results["credentials"] = r.dict()
    except HTTPException as e:
        results["credentials"] = {"error": str(e.detail)}

    return {"results": results}


@router.post("/auto")
def import_auto_endpoint(session: Session = Depends(get_session)):
    """Auto-import all data from hardcoded file paths."""
    import os
    
    # Use hardcoded paths relative to project root
    base_path = Path(__file__).parent.parent  # Go up to gui folder
    enum_file = base_path / "JKP Network Design Draft.xlsx"
    ip_file = base_path / "IP SCHEMA.xlsx"
    credentials_file = base_path / "PHASE 1 CREDENTIALS.xlsx"
    
    results = {}
    raw_results = {}
    try:
        if enum_file.exists():
            raw_results["JKP Network Design Draft.xlsx"] = store_workbook_raw(enum_file, session)
    except Exception as e:
        raw_results["JKP Network Design Draft.xlsx"] = {"error": str(e)}
    try:
        if ip_file.exists():
            raw_results["IP SCHEMA.xlsx"] = store_workbook_raw(ip_file, session)
    except Exception as e:
        raw_results["IP SCHEMA.xlsx"] = {"error": str(e)}
    try:
        if credentials_file.exists():
            raw_results["PHASE 1 CREDENTIALS.xlsx"] = store_workbook_raw(credentials_file, session)
    except Exception as e:
        raw_results["PHASE 1 CREDENTIALS.xlsx"] = {"error": str(e)}

    results["raw"] = raw_results

    # Enum-1 (JKP Network Design Draft)
    try:
        if enum_file.exists():
            r = import_enum1(enum_file, "Enum-1", session)
            results["enum1"] = r.dict()
        else:
            results["enum1"] = {"error": f"File not found: {enum_file}"}
    except HTTPException as e:
        results["enum1"] = {"error": str(e.detail)}
    except Exception as e:
        results["enum1"] = {"error": str(e)}

    # IP Schema - Poles (data extraction from Field Device Details - Poles)
    try:
        if ip_file.exists():
            r = import_ip_schema_data(ip_file, "Field Device Details - Poles", session)
            results["ip_poles"] = r.dict()
        else:
            results["ip_poles"] = {"error": f"File not found: {ip_file}"}
    except HTTPException as e:
        results["ip_poles"] = {"error": str(e.detail)}
    except Exception as e:
        results["ip_poles"] = {"error": str(e)}

    # IP Schema - JBs
    try:
        if ip_file.exists():
            r = import_field_device_jbs(ip_file, "Field Device Details - JB", session)
            results["ip_jbs"] = r.dict()
        else:
            results["ip_jbs"] = {"error": f"File not found: {ip_file}"}
    except HTTPException as e:
        results["ip_jbs"] = {"error": str(e.detail)}
    except Exception as e:
        results["ip_jbs"] = {"error": str(e)}

    # Credentials - Import from all region sheets
    credential_sheets = ['JAMMU', 'SAMBA', 'KATHUA', 'AWANTIPURA', 'BARAMULLA', 'SRINAGAR', 'UDHAMPUR']
    credentials_results = {}
    
    try:
        if credentials_file.exists():
            wb = openpyxl.load_workbook(credentials_file)
            for sheet_name in credential_sheets:
                if sheet_name in wb.sheetnames:
                    try:
                        r = import_credentials(credentials_file, sheet_name, session)
                        credentials_results[sheet_name] = r.dict()
                    except Exception as e:
                        credentials_results[sheet_name] = {"error": str(e)}
            results["credentials"] = credentials_results
        else:
            results["credentials"] = {"error": f"File not found: {credentials_file}"}
    except HTTPException as e:
        results["credentials"] = {"error": str(e.detail)}
    except Exception as e:
        results["credentials"] = {"error": str(e)}

    return {"results": results}


class UploadedFileInfo(BaseModel):
    filename: str
    sheets: list[str]
    detected_type: str


@router.post("/detect-schema")
async def detect_schema(file: UploadFile = File(...)) -> UploadedFileInfo:
    """Upload a file and detect its schema type (enum1, ip-schema, or credentials)"""
    if not file.filename or not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    
    try:
        # Read the uploaded file into memory
        contents = await file.read()
        
        # Create a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(contents)
            tmp_path = Path(tmp_file.name)
        
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=False)
            sheets = wb.sheetnames
            
            # Detect schema type based on sheets and headers
            detected_type = _detect_file_type(tmp_path, sheets)
            
            return UploadedFileInfo(
                filename=file.filename,
                sheets=sheets,
                detected_type=detected_type
            )
        finally:
            # Clean up temp file
            tmp_path.unlink(missing_ok=True)
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to process file: {str(e)}")


def _detect_file_type(path: Path, sheets: list[str]) -> str:
    """Detect file type based on sheets and their content"""
    sheets_lower = [s.lower() for s in sheets]
    
    # Check for JKP Network Design (Enum-1)
    if any("enum" in s for s in sheets_lower):
        return "enum1"
    
    # Check for IP Schema
    if any("field device" in s or "pole" in s or "jb" in s for s in sheets_lower):
        return "ip-schema"
    
    # Check for Credentials
    if any(region.lower() in sheets_lower for region in ['jammu', 'samba', 'kathua', 'awantipura', 'baramulla', 'srinagar', 'udhampur']):
        return "credentials"
    
    # Try to detect by examining actual headers
    for sheet_name in sheets:
        try:
            wb = openpyxl.load_workbook(path, data_only=False)
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            
            # Check first few rows for headers
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if row:
                    headers.extend([str(h).lower() if h else "" for h in row])
            
            headers_str = " ".join(headers)
            
            if "component id" in headers_str or "component type" in headers_str:
                return "enum1"
            if "pole location" in headers_str or "jb id" in headers_str:
                return "ip-schema"
            if "username" in headers_str or "password" in headers_str or "appliance" in headers_str:
                return "credentials"
        except:
            continue
    
    return "unknown"


class FileImportRequest(BaseModel):
    filename: str
    sheet_name: str
    file_content: str  # Base64 encoded
    import_type: str  # "enum1", "ip-schema", or "credentials"


@router.post("/upload-and-import")
async def upload_and_import(
    file: UploadFile = File(...),
    sheet_name: str = "auto",
    import_type: str = "auto",
    session: Session = Depends(get_session)
):
    """Upload a file and import it, auto-detecting type if needed"""
    if not file.filename or not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    
    try:
        # Read the uploaded file
        contents = await file.read()
        
        # Create a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(contents)
            tmp_path = Path(tmp_file.name)
        
        try:
            # Open workbook to get sheets
            wb = openpyxl.load_workbook(tmp_path, data_only=False)
            available_sheets = wb.sheetnames
            raw_info = store_workbook_raw(tmp_path, session, filename_override=file.filename)

            # Auto-detect import type if needed
            if import_type == "auto":
                import_type = _detect_file_type(tmp_path, available_sheets)
            
            if import_type == "unknown":
                raise HTTPException(status_code=400, detail="Could not auto-detect file type. Please specify import_type.")
            
            # Determine sheet to use
            if sheet_name == "auto":
                if import_type == "enum1":
                    sheet_name = "Enum-1" if "Enum-1" in available_sheets else available_sheets[0]
                elif import_type == "ip-schema":
                    sheet_name = "Field Device Details - Poles" if "Field Device Details - Poles" in available_sheets else available_sheets[0]
                elif import_type == "credentials":
                    sheet_name = available_sheets[0]  # Use first sheet for credentials
            
            # Run the appropriate import
            if import_type == "enum1":
                result = import_enum1(tmp_path, sheet_name, session)
            elif import_type == "ip-schema":
                # For IP schema files, try to import both poles and JBs
                result_poles = import_ip_schema_data(tmp_path, sheet_name, session)
                result_jbs_sheet = "Field Device Details - JB" if "Field Device Details - JB" in available_sheets else None
                if result_jbs_sheet:
                    result_jbs = import_field_device_jbs(tmp_path, result_jbs_sheet, session)
                    return {
                        "success": True,
                        "message": "IP Schema imported",
                        "poles": result_poles.dict(),
                        "jbs": result_jbs.dict() if result_jbs_sheet else None
                    }
                return {"success": True, "message": "IP Schema poles imported", "result": result_poles.dict()}
            elif import_type == "credentials":
                result = import_credentials(tmp_path, sheet_name, session)
            else:
                raise HTTPException(status_code=400, detail=f"Unknown import type: {import_type}")
            
            return {
                "success": True,
                "message": f"Successfully imported from {file.filename}",
                "import_type": import_type,
                "sheet": sheet_name,
                "raw": raw_info,
                "result": result.dict()
            }
        finally:
            # Clean up temp file
            tmp_path.unlink(missing_ok=True)
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")


@router.get("/available-sheets")
async def get_available_sheets(file_path: str):
    """Get list of sheets in an Excel file"""
    path = Path(file_path)
    if not path.exists():
        raise HTTPException(status_code=400, detail=f"File not found: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        detected_type = _detect_file_type(path, wb.sheetnames)
        return {"sheets": wb.sheetnames, "detected_type": detected_type}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")

